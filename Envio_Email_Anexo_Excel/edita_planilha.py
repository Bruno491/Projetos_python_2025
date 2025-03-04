from openpyxl import load_workbook
from tabulate import tabulate
import os

def exibir_planilha(planilha):
    tabela = [linha for linha in planilha.iter_rows(values_only=True)]
    print(tabulate(tabela, headers='firstrow', tablefmt='fancy_grid'))

def converter_valor(valor):
    # Tentar converter para número
    try:
        return float(valor)
    except ValueError:
        pass
    
    # Tentar converter para data
    try:
        from datetime import datetime
        return datetime.strptime(valor, "%d/%m/%Y")
    except ValueError:
        pass
    
    # Retornar como string se não for número nem data
    return valor

def edita_planilha(caminho_arquivo, nome_planilha, celula, novo_valor):
    try:
        # Verificar se o arquivo está acessível
        if not os.access(caminho_arquivo, os.W_OK):
            print(f"Permissão negada para editar o arquivo: {caminho_arquivo}")
            return

        # Carregar o arquivo Excel
        workbook = load_workbook(caminho_arquivo)
        
        # Selecionar a planilha desejada
        if nome_planilha in workbook.sheetnames:
            planilha = workbook[nome_planilha]
        else:
            print(f"A planilha '{nome_planilha}' não foi encontrada no arquivo.")
            return
        
        # Exibir a planilha antes da edição
        print("Planilha antes da edição:")
        exibir_planilha(planilha)
        
        # Converter o valor para o tipo apropriado
        novo_valor = converter_valor(novo_valor)
        
        # Editar a célula especificada
        planilha[celula] = novo_valor
        
        # Salvar as alterações no arquivo
        workbook.save(caminho_arquivo)
        print(f"Valor da célula {celula} atualizado para '{novo_valor}' na planilha '{nome_planilha}'.")
        
        # Exibir a planilha após a edição
        print("Planilha após a edição:")
        exibir_planilha(planilha)
    except FileNotFoundError:
        print(f"O arquivo '{caminho_arquivo}' não foi encontrado.")
    except PermissionError:
        print(f"Permissão negada para editar o arquivo: {caminho_arquivo}")
    except Exception as e:
        print(f"Ocorreu um erro ao editar a planilha: {e}")

if __name__ == "__main__":
    caminho_arquivo = input("Informe o caminho do arquivo Excel: ")
    nome_planilha = input("Informe o nome da planilha: ")
    celula = input("Informe a célula que deseja editar (por exemplo, A1): ")
    novo_valor = input("Informe o novo valor para a célula: ")
    
    edita_planilha(caminho_arquivo, nome_planilha, celula, novo_valor)